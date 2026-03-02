"""
Генератор отчётов: Excel и текстовый
"""

import os
import tempfile
from datetime import datetime
from typing import Optional

from data import AUDIT_BLOCKS, STOP_FACTORS


def get_decision_text(pct: float) -> tuple[str, str]:
    if pct >= 80:
        return "GO — ВХОДИТЬ", "Высокий потенциал. Начинайте переговоры немедленно."
    elif pct >= 60:
        return "GO + УСЛОВИЕ", "Потенциал есть. Пилот 90 дней с чёткими KPI."
    elif pct >= 40:
        return "ОСТОРОЖНО", "Серьёзные дыры. Только при жёстких условиях договора."
    else:
        return "NO GO", "Высокий риск. Отказ или повторный аудит через 6 месяцев."


async def generate_excel_report(audit_id: str, audit: dict, storage) -> str:
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise Exception("openpyxl не установлен")

    wb = openpyxl.Workbook()

    # ── COLOR PALETTE ──────────────────────────────────────────────────────
    C_BRAND   = "1A3C5E"
    C_ACCENT  = "E8734A"
    C_GREEN   = "2D7A4F"
    C_RED     = "B03A2E"
    C_YELLOW  = "F59E0B"
    C_LIGHT   = "EDF2F7"
    C_LORANGE = "FDF0EB"
    C_LGREEN  = "EDFAF4"
    C_WHITE   = "FFFFFF"
    C_DARK    = "1A202C"
    C_GRAY    = "718096"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color=C_DARK, size=11, italic=False):
        return Font(bold=bold, color=color, size=size, name="Arial", italic=italic)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style='thin', color='CBD5E0')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    company = audit.get('company', 'Компания')
    created = datetime.fromisoformat(audit['created_at']).strftime('%d.%m.%Y')
    total, max_total = storage.get_total_score(audit_id)
    pct = round(total / max_total * 100) if max_total > 0 else 0
    decision, advice = get_decision_text(pct)

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: ИТОГОВЫЙ ДАШБОРД
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "📊 Итоги"

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22

    # HEADER
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = "АУДИТ БИЗНЕСА ПЕРЕД ПАРТНЁРСКИМ ВХОДОМ"
    c.fill = fill(C_BRAND)
    c.font = font(bold=True, color=C_WHITE, size=14)
    c.alignment = align("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = f"Компания: {company}   |   Дата: {created}   |   Аудитор: _______________"
    c.fill = fill(C_LIGHT)
    c.font = font(color=C_DARK, size=11)
    c.alignment = align("center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    # BLOCK RESULTS TABLE header
    headers = ["БЛОК АУДИТА", "МАКС.", "НАБРАНО", "% ВЫПОЛН.", "ПРОГРЕСС", "СТАТУС"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.fill = fill(C_BRAND)
        c.font = font(bold=True, color=C_WHITE, size=10)
        c.alignment = align("center")
        c.border = thin_border
    ws.row_dimensions[header_row].height = 22

    row = 5
    for i, block in enumerate(AUDIT_BLOCKS):
        answers = storage.get_block_answers(audit_id, i)
        block_score = sum(a.get('score', 0) for a in answers.values())
        block_pct = round(block_score / block['max'] * 100) if block['max'] > 0 else 0
        answered = len(answers)
        total_q = len(block['questions'])

        # Color by performance
        if block_pct >= 70:
            row_color = C_LGREEN
            status = "✅ Хорошо"
        elif block_pct >= 50:
            row_color = "FFFBEA"
            status = "🟡 Средне"
        elif block_pct >= 30 and answered > 0:
            row_color = C_LORANGE
            status = "⚠️ Слабо"
        elif answered == 0:
            row_color = C_LIGHT
            status = "⬜ Не заполнен"
        else:
            row_color = "FDECEA"
            status = "🔴 Критично"

        bar_len = round(block_pct / 10)
        bar = "█" * bar_len + "░" * (10 - bar_len)

        cells_data = [
            (block['title'].replace('📊 ', '').replace('💼 ', '').replace('📣 ', '').replace('⚙️ ', '').replace('👤 ', '').replace('🌍 ', ''), "left"),
            (block['max'], "center"),
            (block_score, "center"),
            (f"{block_pct}%", "center"),
            (f"{bar} {answered}/{total_q}", "left"),
            (status, "center"),
        ]
        for col, (val, h_align) in enumerate(cells_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill(row_color)
            c.font = font(bold=(col == 1), color=C_DARK, size=10)
            c.alignment = align(h_align)
            c.border = thin_border
        ws.row_dimensions[row].height = 20
        row += 1

    # TOTAL ROW
    for col, val in enumerate(["ИТОГО", max_total, total, f"{pct}%", "", ""], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(C_ACCENT)
        c.font = font(bold=True, color=C_WHITE, size=11)
        c.alignment = align("center" if col > 1 else "left")
        c.border = thin_border
    ws.row_dimensions[row].height = 24
    row += 2

    # DECISION BOX
    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    c.value = "РЕШЕНИЕ О ВХОДЕ В БИЗНЕС"
    c.fill = fill(C_BRAND)
    c.font = font(bold=True, color=C_WHITE, size=12)
    c.alignment = align("center")
    ws.row_dimensions[row].height = 24
    row += 1

    if pct >= 80:
        dec_color = "2D7A4F"
        dec_bg = C_LGREEN
    elif pct >= 60:
        dec_color = "C47F1A"
        dec_bg = "FFFBEA"
    elif pct >= 40:
        dec_color = "D97706"
        dec_bg = C_LORANGE
    else:
        dec_color = C_RED
        dec_bg = "FDECEA"

    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    c.value = f"  {decision}  —  {pct}% из 100%"
    c.fill = fill(dec_bg)
    c.font = Font(bold=True, color=dec_color, size=14, name="Arial")
    c.alignment = align("center")
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    c.value = f"  {advice}"
    c.fill = fill(dec_bg)
    c.font = font(italic=True, color=C_DARK, size=11)
    c.alignment = align("center")
    ws.row_dimensions[row].height = 22
    row += 2

    # DECISION MATRIX
    matrix_header = ws.cell(row=row, column=1, value="ШКАЛА ПРИНЯТИЯ РЕШЕНИЙ")
    ws.merge_cells(f'A{row}:F{row}')
    matrix_header.fill = fill(C_BRAND)
    matrix_header.font = font(bold=True, color=C_WHITE, size=10)
    matrix_header.alignment = align("center")
    ws.row_dimensions[row].height = 20
    row += 1

    matrix_data = [
        ("480–600 баллов (80–100%)", "GO — ВХОДИТЬ", "Высокий потенциал. Переговоры немедленно", C_LGREEN),
        ("360–479 баллов (60–79%)", "GO + УСЛОВИЕ", "Пилот 90 дней с чёткими KPI", "FFFBEA"),
        ("240–359 баллов (40–59%)", "ОСТОРОЖНО", "Только при жёстких условиях договора", C_LORANGE),
        ("0–239 баллов (<40%)", "NO GO", "Отказ или повторный аудит через 6 мес.", "FDECEA"),
    ]

    for pts, dec, note, bg in matrix_data:
        for col, (val, w) in enumerate([(pts, 16), (dec, 12), (note, 20)], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill(bg)
            c.font = font(bold=(col == 2), color=C_DARK, size=10)
            c.alignment = align("left", wrap=True)
            c.border = thin_border
        ws.merge_cells(f'D{row}:F{row}')
        ws.row_dimensions[row].height = 18
        row += 1

    # ══════════════════════════════════════════════════════════════════════
    # SHEETS 2–7: ДЕТАЛЬНЫЕ ОТВЕТЫ ПО КАЖДОМУ БЛОКУ
    # ══════════════════════════════════════════════════════════════════════
    score_colors = {
        0: "FDECEA", 1: "FEE2D5", 2: C_LORANGE,
        3: "FFFBEA", 4: C_LGREEN, 5: "D1FAE5"
    }
    score_labels = {
        0: "❌ Нет/Критично", 1: "⚠️ Очень слабо", 2: "😐 Слабо",
        3: "🆗 Средне", 4: "👍 Хорошо", 5: "⭐ Отлично"
    }

    for block_idx, block in enumerate(AUDIT_BLOCKS):
        answers = storage.get_block_answers(audit_id, block_idx)
        block_score = sum(a.get('score', 0) for a in answers.values())
        block_pct = round(block_score / block['max'] * 100) if block['max'] > 0 else 0

        sheet_name = block['short']
        ws2 = wb.create_sheet(title=sheet_name)

        # Column widths
        ws2.column_dimensions['A'].width = 4
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 38
        ws2.column_dimensions['D'].width = 8
        ws2.column_dimensions['E'].width = 16
        ws2.column_dimensions['F'].width = 35

        # Block header
        ws2.merge_cells('A1:F1')
        c = ws2['A1']
        title = block['title'].replace('📊 ', '').replace('💼 ', '').replace('📣 ', '').replace('⚙️ ', '').replace('👤 ', '').replace('🌍 ', '')
        c.value = title.upper()
        c.fill = fill(C_BRAND)
        c.font = font(bold=True, color=C_WHITE, size=12)
        c.alignment = align("center")
        ws2.row_dimensions[1].height = 26

        ws2.merge_cells('A2:F2')
        c = ws2['A2']
        c.value = f"Набрано: {block_score} / {block['max']} баллов ({block_pct}%)   |   {block['description']}"
        c.fill = fill(C_LIGHT)
        c.font = font(color=C_DARK, size=10, italic=True)
        c.alignment = align("center")
        ws2.row_dimensions[2].height = 18

        ws2.row_dimensions[3].height = 6

        # Table header
        col_headers = ["#", "ПУНКТ АУДИТА", "КАК ПРОВЕРИТЬ", "0–5", "ОЦЕНКА", "КОММЕНТАРИЙ / ФАКТЫ"]
        for col, h in enumerate(col_headers, 1):
            c = ws2.cell(row=4, column=col, value=h)
            c.fill = fill(C_BRAND)
            c.font = font(bold=True, color=C_WHITE, size=9)
            c.alignment = align("center")
            c.border = thin_border
        ws2.row_dimensions[4].height = 20

        for q_idx, q in enumerate(block['questions']):
            row_num = 5 + q_idx
            answer = answers.get(str(q_idx), {})
            score = answer.get('score', None)
            comment = answer.get('comment', '')

            if score is not None:
                row_bg = score_colors.get(score, C_WHITE)
                score_label = score_labels.get(score, str(score))
                score_val = score
            else:
                row_bg = "F7FAFC"
                score_label = "— Не заполнено"
                score_val = "—"

            row_data = [
                (q_idx + 1, "center", False),
                (q['title'], "left", True),
                (q['how'], "left", False),
                (score_val, "center", True),
                (score_label, "center", False),
                (comment, "left", False),
            ]

            for col, (val, h_align, bold) in enumerate(row_data, 1):
                c = ws2.cell(row=row_num, column=col, value=val)
                c.fill = fill(row_bg)
                c.font = font(bold=bold, color=C_DARK, size=9,
                             italic=(col == 3))
                c.alignment = align(h_align, wrap=True)
                c.border = thin_border

            ws2.row_dimensions[row_num].height = 36

        # Totals row
        total_row = 5 + len(block['questions'])
        ws2.merge_cells(f'A{total_row}:C{total_row}')
        c = ws2.cell(row=total_row, column=1, value=f"ИТОГО: {block_score} / {block['max']} баллов")
        c.fill = fill(C_ACCENT)
        c.font = font(bold=True, color=C_WHITE, size=10)
        c.alignment = align("right")
        c.border = thin_border
        ws2.cell(row=total_row, column=4, value=f"{block_pct}%").fill = fill(C_ACCENT)
        ws2.cell(row=total_row, column=4).font = font(bold=True, color=C_WHITE, size=10)
        ws2.cell(row=total_row, column=4).alignment = align("center")
        ws2.cell(row=total_row, column=4).border = thin_border
        ws2.merge_cells(f'E{total_row}:F{total_row}')
        ws2.cell(row=total_row, column=5).fill = fill(C_ACCENT)
        ws2.cell(row=total_row, column=5).border = thin_border
        ws2.row_dimensions[total_row].height = 22

    # ══════════════════════════════════════════════════════════════════════
    # SHEET: СТОП-ФАКТОРЫ
    # ══════════════════════════════════════════════════════════════════════
    ws_stop = wb.create_sheet(title="🚩 Стоп-факторы")
    ws_stop.column_dimensions['A'].width = 5
    ws_stop.column_dimensions['B'].width = 55
    ws_stop.column_dimensions['C'].width = 20

    ws_stop.merge_cells('A1:C1')
    c = ws_stop['A1']
    c.value = "🚩 СТОП-ФАКТОРЫ — АВТОМАТИЧЕСКИЙ ОТКАЗ"
    c.fill = fill(C_RED)
    c.font = font(bold=True, color=C_WHITE, size=12)
    c.alignment = align("center")
    ws_stop.row_dimensions[1].height = 26

    ws_stop.merge_cells('A2:C2')
    c = ws_stop['A2']
    c.value = "Один подтверждённый стоп-фактор = NO GO без исключений"
    c.fill = fill("FDECEA")
    c.font = font(italic=True, color=C_RED, size=10)
    c.alignment = align("center")
    ws_stop.row_dimensions[2].height = 18

    for col, h in enumerate(["#", "СТОП-ФАКТОР", "ВЫЯВЛЕН?"], 1):
        c = ws_stop.cell(row=3, column=col, value=h)
        c.fill = fill(C_RED)
        c.font = font(bold=True, color=C_WHITE, size=10)
        c.alignment = align("center")
        c.border = thin_border
    ws_stop.row_dimensions[3].height = 20

    from data import STOP_FACTORS
    for i, factor in enumerate(STOP_FACTORS):
        row_num = 4 + i
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        ws_stop.cell(row=row_num, column=1, value=i + 1).fill = fill(bg)
        ws_stop.cell(row=row_num, column=1).alignment = align("center")
        ws_stop.cell(row=row_num, column=1).border = thin_border
        ws_stop.cell(row=row_num, column=1).font = font(size=9)

        ws_stop.cell(row=row_num, column=2, value=factor).fill = fill(bg)
        ws_stop.cell(row=row_num, column=2).alignment = align(wrap=True)
        ws_stop.cell(row=row_num, column=2).border = thin_border
        ws_stop.cell(row=row_num, column=2).font = font(size=9)

        ws_stop.cell(row=row_num, column=3, value="☐  Да  /  ☐  Нет").fill = fill(bg)
        ws_stop.cell(row=row_num, column=3).alignment = align("center")
        ws_stop.cell(row=row_num, column=3).border = thin_border
        ws_stop.cell(row=row_num, column=3).font = font(size=9)
        ws_stop.row_dimensions[row_num].height = 22

    # ══════════════════════════════════════════════════════════════════════
    # SHEET: БЫСТРЫЕ ПОБЕДЫ
    # ══════════════════════════════════════════════════════════════════════
    ws_qw = wb.create_sheet(title="⚡ Быстрые победы")
    ws_qw.column_dimensions['A'].width = 28
    ws_qw.column_dimensions['B'].width = 35
    ws_qw.column_dimensions['C'].width = 20
    ws_qw.column_dimensions['D'].width = 12
    ws_qw.column_dimensions['E'].width = 10

    ws_qw.merge_cells('A1:E1')
    c = ws_qw['A1']
    c.value = "⚡ МАТРИЦА БЫСТРЫХ ПОБЕД — ПЕРВЫЕ 30 ДНЕЙ"
    c.fill = fill("2D7A4F")
    c.font = font(bold=True, color=C_WHITE, size=12)
    c.alignment = align("center")
    ws_qw.row_dimensions[1].height = 26

    quick_wins = [
        ("Запустить AmoCRM (базово)", "Воронка + перенос базы + обучение 2 часа", "+15–20% к конверсии", "7 дней"),
        ("Скрипт входящего звонка", "Написать, записать образец, внедрить в команду", "+20–30% к входящим", "5 дней"),
        ("Дожим «потерянных» за 6 мес.", "3 касания всем, кто сказал «подумаю»", "+10–20% к выручке", "3 дня"),
        ("Апсейл при каждой продаже", "Скрипт: предложить пакет выше / доп.опции", "+15–25% к ср.чеку", "7 дней"),
        ("Убрать убыточные продукты", "Анализ маржи → стоп-продажа минусовых", "Маржа +3–5%", "3 дня"),
        ("Запустить 1 новый рекламный канал", "Яндекс Директ или таргет ВК если не было", "+30–50% лидов", "14 дней"),
        ("Еженедельная планёрка по KPI", "15 мин в пн. с командой: цифры + задачи", "Фокус и контроль", "День 1"),
        ("Запрос отзывов у клиентов", "Скрипт запроса + ссылка на Яндекс.Карты", "+0.3–0.5 рейтинга", "7 дней"),
        ("P&L таблица еженедельная", "Google Таблица: выручка, маржа, расходы", "Видимость финансов", "3 дня"),
        ("Правило: каждая сделка в CRM", "Нет записи = нет бонуса. С первого дня", "100% данных", "День 1"),
    ]

    for col, h in enumerate(["ДЕЙСТВИЕ", "КАК СДЕЛАТЬ", "ЭФФЕКТ", "СРОК", "✓"], 1):
        c = ws_qw.cell(row=2, column=col, value=h)
        c.fill = fill("2D7A4F")
        c.font = font(bold=True, color=C_WHITE, size=9)
        c.alignment = align("center")
        c.border = thin_border
    ws_qw.row_dimensions[2].height = 20

    for i, (action, how, effect, deadline) in enumerate(quick_wins):
        row_num = 3 + i
        bg = C_LGREEN if i % 2 == 0 else C_WHITE
        for col, (val, bold) in enumerate([(action, True), (how, False), (effect, True), (deadline, False), ("☐", False)], 1):
            c = ws_qw.cell(row=row_num, column=col, value=val)
            c.fill = fill(bg)
            c.font = font(bold=bold and col < 4, color="2D7A4F" if col in [1, 3] else C_DARK, size=9)
            c.alignment = align("center" if col in [3, 4, 5] else "left", wrap=True)
            c.border = thin_border
        ws_qw.row_dimensions[row_num].height = 24

    # Save to temp file
    tmp = tempfile.mktemp(suffix='.xlsx')
    wb.save(tmp)
    return tmp


def generate_text_report(audit_id: str, audit: dict, storage) -> str:
    """Fallback текстовый отчёт"""
    company = audit.get('company', 'Компания')
    total, max_total = storage.get_total_score(audit_id)
    pct = round(total / max_total * 100) if max_total > 0 else 0
    decision, advice = get_decision_text(pct)

    lines = [
        f"<b>АУДИТ БИЗНЕСА: {company}</b>",
        f"Дата: {datetime.now().strftime('%d.%m.%Y')}",
        "",
        f"<b>РЕЗУЛЬТАТЫ ПО БЛОКАМ:</b>",
    ]

    for i, block in enumerate(AUDIT_BLOCKS):
        answers = storage.get_block_answers(audit_id, i)
        block_score = sum(a.get('score', 0) for a in answers.values())
        block_pct = round(block_score / block['max'] * 100) if block['max'] > 0 else 0
        short = block['short']
        lines.append(f"• {short}: {block_score}/{block['max']} ({block_pct}%)")

    lines.extend([
        "",
        f"<b>ИТОГО: {total}/{max_total} ({pct}%)</b>",
        f"<b>Решение: {decision}</b>",
        f"{advice}",
    ])

    return "\n".join(lines)
